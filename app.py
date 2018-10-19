#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from xlrd import open_workbook

import ipdb, re, sys

def strip_eng(txt):
    pattern = r'^[\x00-\x7F]+'
    return re.sub(pattern, '', txt)

def translate_sheet(current_wb):
    print '=== loading workbook ' + str(current_wb)
    ger_wb = load_workbook(str(current_wb) + '.xlsx')
    ans_wb = open_workbook('answered.xls')

    ger_ws = ger_wb.active

    if ger_ws.title != str(current_wb):
        ger_ws.title = str(current_wb)

    print '=== listing sheets in answered.xls'
    for n in ans_wb.sheet_names():
        print n

    print

    ans_ws = ans_wb.sheet_by_name(u'试题' + str(current_wb))

    print '=== reading ' + ans_ws.name

    # answsers start from row 2 in answered sheet, 
    for i in range(1, ans_ws.nrows):
        print 'translating row {}'.format(i)
        for j, c in enumerate(ans_ws.row(i)):
            j = j + 1
            ger_ws_value = unicode(ger_ws.cell(i, j).value)
            if not ger_ws_value:
                ger_ws_value = ''
            else:
                ger_ws_value = ger_ws_value.strip() + u'\n'
            if j <= 4:
                try:
                    chn_txt = strip_eng(unicode(c.value))
                    ger_ws_value = ger_ws_value + chn_txt
                except TypeError as e:
                    print '!!! failed to translate', c.value, type(c.value)
            else:
                ger_ws_value = c.value
            ger_ws.cell(i, j).value = ger_ws_value

    ger_wb.save('{}_translated.xlsx'.format(current_wb))

if __name__ == '__main__':
    if len(sys.argv) != 2:
        print "Usage: {} filenumber".format(sys.argv[0])
        sys.exit(1)
    translate_sheet(sys.argv[1])

