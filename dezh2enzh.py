#!/usr/bin/env python
# -*- coding: utf-8 -*-

TOTAL = 37

from openpyxl import load_workbook, Workbook
from xlrd import open_workbook

import ipdb, re, sys

def is_zh(ch):
    try:
        return ch >= u'\u4e00'
    except:
        return False

def strip_zh(txt):
    for i in range(len(txt)):
        if is_zh(txt[i]):
            return txt[:i]
    return txt

def strip_de(txt):
    for i in range(len(txt)):
        if is_zh(txt[i]):
            return txt[i:]
    return ''

def transpose(workbook, i, dezh_sheet, enzh_sheet):
    new_sheet = None
    if i == 0:
        workbook.worksheets[0].title = str(1)
        new_sheet = workbook.worksheets[0]
    else:
        new_sheet = workbook.create_sheet(str(i+1))

    for row in enzh_sheet.iter_rows():
        for cell in row:
            if cell.row == 1:
                new_sheet.cell(cell.row, cell.col_idx).value = cell.value
            else:
                if cell.value:
                    new_sheet.cell(cell.row, cell.col_idx).value = strip_zh(unicode(cell.value)) + '\n' + strip_de(unicode(dezh_sheet.cell_value(cell.row - 1, cell.col_idx - 1)))



if __name__ == '__main__':
    result = Workbook()
    for i in range(TOTAL):
        try:
            print "loading", 'engchn/test{}.xlsx'.format(i+1), "...",
            enzh_workbook = load_workbook('engchn/test{}.xlsx'.format(i+1))
            print "done"
        except:
            print "could not read 'engchn/test{}.xlsx'".format(i+1)
            continue
        try:
            dezh_workbook = open_workbook(u'gerchn/百应批导模板_德中文_{}.xls'.format(i+1))
        except:
            print "could not read gerchn/百应批导模板_德中文_{}.xls".format(i+1)
            continue

        enzh = enzh_workbook.worksheets[0]
        dezh = dezh_workbook.sheet_by_index(0)
        transpose(result, i, dezh, enzh)
    result.save('new_engchn.xlsx')

#ipdb.set_trace()
