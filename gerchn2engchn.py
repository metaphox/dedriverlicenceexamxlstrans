#!/usr/bin/env python
# -*- coding: utf-8 -*-

TOTAL = 37

from openpyxl import load_workbook, Workbook
from xlrd import open_workbook

import ipdb, re, sys

def is_zh(ch):
    try:
        return ord(ch) > 0x4e00
    except:
        return False

def transpose(workbook, i, dezh_sheet, enzh_sheet):
    new_sheet = None
    if i == 0:
        workbook.worksheets[0].title = str(1)
        new_sheet = workbook.worksheets[0]
    else:
        new_sheet = workbook.create_sheet(str(i+1))
    print enzh_sheet['A2'].value, '\n', dezh_sheet.cell_value(1, 0)


if __name__ == '__main__':
    result = Workbook()
    for i in range(TOTAL):
        try:
            enzh_workbook = load_workbook('engchn/test{}.xlsx'.format(i+1))
        except:
            print "could not read 'engchn/test{}.xlsx'".format(i+1)
            continue
        try:
            dezh_workbook = open_workbook(u'gerchn/百应批导模板_德中文_{}.xls'.format(i+1))
        except:
            print "could not read gerchn/百应批导模板_德中文_{}.xls".format(i+1)
            continue

        enzh = enzh_workbook.worksheets[0]
        dezh = dezh_workbook.sheet_by_index(0)
        transpose(result, i, dezh, enzh)
    result.save('new_engchn.xlsx')

#ipdb.set_trace()
